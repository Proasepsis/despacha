from pathlib import Path

from django.test import SimpleTestCase
from openpyxl import Workbook

from core.adaptadores.plantilla.adaptador import AdaptadorPlantilla


ENCABEZADOS = [
    "NÚMERO DE DOCUMENTO",
    "TIPO DE COMPROBANTE",
    "CÓDIGO COMPROBANTE",
    "CUENTA CONTABLE",
    "DÉBITO O CRÉDITO",
    "LÍNEA PRODUCTO",
    "GRUPO PRODUCTO",
    "CÓDIGO PRODUCTO",
    "CANTIDAD",
    "LOTE",
    "NIT",
    "CÓDIGO DE LA CIUDAD",
    "DESCRIPCIÓN DE LA SECUENCIA",
]


class LecturaLoteStringTest(SimpleTestCase):
    def test_lote_con_punto_final_se_preserva_al_leer_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"

        for col, encabezado in enumerate(ENCABEZADOS, start=1):
            ws.cell(row=5, column=col, value=encabezado)

        fila = ["DOC001", "F", "1", "143505", "C", "150", "0005", "000005", "10",
                "120010925.", "800000", "11001", "Desc"]
        for j, valor in enumerate(fila, start=1):
            ws.cell(row=6, column=j, value=valor)

        ruta = Path("/tmp/test_lote_punto_integracion.xlsx")
        wb.save(ruta)

        adaptador = AdaptadorPlantilla()
        documentos = adaptador.parse(ruta)
        ruta.unlink(missing_ok=True)

        linea = documentos[0].lineas[0]
        self.assertEqual(linea.lote_raw, "120010925.",
                         "El lote con punto final debe preservarse al leer el Excel")

    def test_lote_con_comilla_slash_espacios_se_preserva_crudo(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"

        for col, encabezado in enumerate(ENCABEZADOS, start=1):
            ws.cell(row=5, column=col, value=encabezado)

        fila = ["DOC001", "F", "1", "143505", "C", "150", "0005", "000005", "10",
                "'15F22/2579", "800000", "11001", "Desc"]
        for j, valor in enumerate(fila, start=1):
            ws.cell(row=6, column=j, value=valor)

        ruta = Path("/tmp/test_lote_crudo_integracion.xlsx")
        wb.save(ruta)

        adaptador = AdaptadorPlantilla()
        documentos = adaptador.parse(ruta)
        ruta.unlink(missing_ok=True)

        linea = documentos[0].lineas[0]
        self.assertEqual(linea.lote_raw, "'15F22/2579")

    def test_lote_con_espacios_internos_se_preserva_crudo(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"

        for col, encabezado in enumerate(ENCABEZADOS, start=1):
            ws.cell(row=5, column=col, value=encabezado)

        fila = ["DOC001", "F", "1", "143505", "C", "150", "0005", "000005", "10",
                "1 5L 23", "800000", "11001", "Desc"]
        for j, valor in enumerate(fila, start=1):
            ws.cell(row=6, column=j, value=valor)

        ruta = Path("/tmp/test_lote_espacios_integracion.xlsx")
        wb.save(ruta)

        adaptador = AdaptadorPlantilla()
        documentos = adaptador.parse(ruta)
        ruta.unlink(missing_ok=True)

        linea = documentos[0].lineas[0]
        self.assertEqual(linea.lote_raw, "1 5L 23")
