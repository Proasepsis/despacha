from pathlib import Path
from decimal import Decimal

from django.test import TestCase
from openpyxl import Workbook

from core.adaptadores.registry import obtener, disponibles
from core.adaptadores.plantilla.adaptador import AdaptadorPlantilla


ENCABEZADOS = [
    "NÚMERO DE DOCUMENTO",
    "TIPO DE COMPROBANTE",
    "CÓDIGO COMPROBANTE",
    "CUENTA CONTABLE",
    "DÉBITO O CRÉDITO",
    "LÍNEA PRODUCTO",
    "GRUPO PRODUCTO",
    "CÓDIGO PRODUCTO",
    "CANTIDAD",
    "LOTE",
    "NIT",
    "CÓDIGO DE LA CIUDAD",
    "DESCRIPCIÓN DE LA SECUENCIA",
]


def _crear_excel_plantilla(ruta: Path, filas: list[list]) -> None:
    """Crea un archivo Excel con estructura PLANTILLA: encabezados en fila 5, datos desde fila 6."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    for col, encabezado in enumerate(ENCABEZADOS, start=1):
        ws.cell(row=5, column=col, value=encabezado)

    for i, fila in enumerate(filas, start=6):
        for j, valor in enumerate(fila, start=1):
            ws.cell(row=i, column=j, value=valor)

    wb.save(ruta)


class AdaptadorPlantillaFiltroTest(TestCase):
    def setUp(self):
        self.adaptador = AdaptadorPlantilla()

    def test_registro_disponible(self):
        adaptador = obtener("PLANTILLA")
        self.assertEqual(adaptador.nombre, "PLANTILLA")
        self.assertIn("PLANTILLA", disponibles())

    def test_validacion_falta_hoja1(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "OtraHoja"
        ruta = Path("/tmp/test_falta_hoja1.xlsx")
        wb.save(ruta)

        with self.assertRaises(ValueError) as ctx:
            self.adaptador.validar(ruta)
        self.assertIn("Hoja1", str(ctx.exception))

        ruta.unlink(missing_ok=True)

    def test_validacion_encabezados_faltantes(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        # Solo ponemos algunos encabezados (faltan varios)
        ws.cell(row=5, column=1, value="NÚMERO DE DOCUMENTO")
        ws.cell(row=5, column=2, value="OTRA COSA")
        ruta = Path("/tmp/test_falta_encabezados.xlsx")
        wb.save(ruta)

        with self.assertRaises(ValueError) as ctx:
            self.adaptador.validar(ruta)
        self.assertIn("fila 5", str(ctx.exception))

        ruta.unlink(missing_ok=True)

    def test_filtro_criterios(self):
        filas = [
            # Fila válida: F+1+14xx+C
            ["DOC001", "F", "1", "143505", "C", "150", "0005", "000005", "10", "15F22", "800000", "11001", "Desc A"],
            # Fila válida: H+5+14xx+C
            ["DOC002", "H", "5", "143510", "C", "151", "0010", "000010", "5", "15G33", "800001", "11001", "Desc B"],
            # Fila válida: S+1+14xx+C
            ["DOC003", "S", "1", "143515", "C", "152", "0015", "000015", "3", "15H44", "800002", "11001", "Desc C"],
            # Fila válida: T+10+14xx+C
            ["DOC004", "T", "10", "143520", "C", "153", "0020", "000020", "7", "15I55", "800003", "11001", "Desc D"],
            # Fila inválida: tipo Z no permitido
            ["DOC005", "Z", "1", "143525", "C", "154", "0025", "000025", "2", "15J66", "800004", "11001", "Desc E"],
            # Fila inválida: F pero código no es 1
            ["DOC006", "F", "5", "143530", "C", "155", "0030", "000030", "4", "15K77", "800005", "11001", "Desc F"],
            # Fila inválida: cuenta no empieza por 14
            ["DOC007", "F", "1", "413535", "C", "156", "0035", "000035", "1", "15L88", "800006", "11001", "Desc G"],
            # Fila inválida: débito en vez de crédito
            ["DOC008", "F", "1", "143540", "D", "157", "0040", "000040", "6", "15M99", "800007", "11001", "Desc H"],
            # Fila válida: S+1+14xx+C
            ["DOC009", "S", "1", "143545", "C", "158", "0045", "000045", "8", "15N00", "800008", "11001", "Desc I"],
            # Fila inválida: H pero código no es 5
            ["DOC010", "H", "1", "143550", "C", "159", "0050", "000050", "9", "2A3B4C", "800009", "11001", "Desc J"],
        ]

        ruta = Path("/tmp/test_filtro_criterios.xlsx")
        _crear_excel_plantilla(ruta, filas)

        documentos = self.adaptador.parse(ruta)
        ruta.unlink(missing_ok=True)

        facturas = [d.factura for d in documentos]
        esperadas = ["DOC001", "DOC002", "DOC003", "DOC004", "DOC009"]
        self.assertEqual(sorted(facturas), sorted(esperadas))

        total_lineas = sum(len(d.lineas) for d in documentos)
        self.assertEqual(total_lineas, 5)

    def test_agrupacion_por_documento(self):
        filas = [
            ["DOC001", "F", "1", "143505", "C", "150", "0005", "000005", "10", "L1", "800000", "11001", "Desc 1"],
            ["DOC001", "F", "1", "143510", "C", "151", "0006", "000006", "5", "L2", "800000", "11001", "Desc 2"],
            ["DOC001", "F", "1", "143515", "C", "152", "0007", "000007", "3", "L3", "800000", "11001", "Desc 3"],
            ["DOC001", "F", "1", "413520", "D", "153", "0008", "000008", "2", "L4", "800000", "11001", "Desc 4"],  # inválida
        ]

        ruta = Path("/tmp/test_agrupacion.xlsx")
        _crear_excel_plantilla(ruta, filas)

        documentos = self.adaptador.parse(ruta)
        ruta.unlink(missing_ok=True)

        self.assertEqual(len(documentos), 1)
        doc = documentos[0]
        self.assertEqual(doc.factura, "DOC001")
        self.assertEqual(len(doc.lineas), 3)

    def test_preservacion_lote_string(self):
        filas = [
            ["DOC001", "F", "1", "143505", "C", "150", "0005", "000005", "10", "120010925.", "800000", "11001", "Desc"],
        ]

        ruta = Path("/tmp/test_lote_string.xlsx")
        _crear_excel_plantilla(ruta, filas)

        documentos = self.adaptador.parse(ruta)
        ruta.unlink(missing_ok=True)

        linea = documentos[0].lineas[0]
        self.assertEqual(linea.lote_raw, "120010925.")

    def test_lote_sin_punto_no_se_agrega(self):
        filas = [
            ["DOC001", "F", "1", "143505", "C", "150", "0005", "000005", "10", "120010925", "800000", "11001", "Desc"],
        ]

        ruta = Path("/tmp/test_lote_sin_punto.xlsx")
        _crear_excel_plantilla(ruta, filas)

        documentos = self.adaptador.parse(ruta)
        ruta.unlink(missing_ok=True)

        linea = documentos[0].lineas[0]
        self.assertEqual(linea.lote_raw, "120010925")

    def test_lote_con_comilla_se_preserva(self):
        filas = [
            ["DOC001", "F", "1", "143505", "C", "150", "0005", "000005", "10", "'121570226", "800000", "11001", "Desc"],
        ]

        ruta = Path("/tmp/test_lote_comilla.xlsx")
        _crear_excel_plantilla(ruta, filas)

        documentos = self.adaptador.parse(ruta)
        ruta.unlink(missing_ok=True)

        linea = documentos[0].lineas[0]
        self.assertEqual(linea.lote_raw, "'121570226")

    def test_cantidad_decimal(self):
        filas = [
            ["DOC001", "F", "1", "143505", "C", "150", "0005", "000005", "7.5", "L1", "800000", "11001", "Desc"],
        ]

        ruta = Path("/tmp/test_cantidad.xlsx")
        _crear_excel_plantilla(ruta, filas)

        documentos = self.adaptador.parse(ruta)
        ruta.unlink(missing_ok=True)

        linea = documentos[0].lineas[0]
        self.assertEqual(linea.cantidad_origen, Decimal("7.5"))

    def test_producto_codigo_armado(self):
        filas = [
            ["DOC001", "F", "1", "143505", "C", "150", "0005", "000005", "10", "L1", "800000", "11001", "Desc"],
        ]

        ruta = Path("/tmp/test_codigo_armado.xlsx")
        _crear_excel_plantilla(ruta, filas)

        documentos = self.adaptador.parse(ruta)
        ruta.unlink(missing_ok=True)

        linea = documentos[0].lineas[0]
        self.assertEqual(linea.producto_codigo, "1500005000005")

    def test_armar_codigo_producto_varios_casos(self):
        from core.adaptadores.plantilla.adaptador import _armar_codigo_producto

        self.assertEqual(_armar_codigo_producto(150, 5, 5), "1500005000005")
        self.assertEqual(_armar_codigo_producto(150.0, 5.0, 5.0), "1500005000005")
        self.assertEqual(_armar_codigo_producto("150", "0005", "000005"), "1500005000005")
        self.assertEqual(_armar_codigo_producto(1, 1, 1), "0010001000001")

    def test_armar_codigo_producto_partes_vacias(self):
        from core.adaptadores.plantilla.adaptador import _armar_codigo_producto

        with self.assertRaises(ValueError):
            _armar_codigo_producto("", "0005", "000005")

    def test_archivo_vacio_sin_datos(self):
        ruta = Path("/tmp/test_vacio.xlsx")
        _crear_excel_plantilla(ruta, [])

        documentos = self.adaptador.parse(ruta)
        ruta.unlink(missing_ok=True)

        self.assertEqual(len(documentos), 0)

    def test_nit_y_ciudad_se_propagan(self):
        filas = [
            ["DOC001", "F", "1", "143505", "C", "150", "0005", "000005", "10", "L1", "900123456", "11001", "Desc 1"],
            ["DOC001", "F", "1", "143510", "C", "151", "0006", "000006", "5", "L2", "900123456", "11001", "Desc 2"],
        ]

        ruta = Path("/tmp/test_nit_ciudad.xlsx")
        _crear_excel_plantilla(ruta, filas)

        documentos = self.adaptador.parse(ruta)
        ruta.unlink(missing_ok=True)

        doc = documentos[0]
        self.assertEqual(doc.nit, "900123456")
        self.assertEqual(doc.codigo_ciudad, "11001")
