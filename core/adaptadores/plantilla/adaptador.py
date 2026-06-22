from pathlib import Path
from decimal import Decimal

from openpyxl import load_workbook

from ..base import AdaptadorFormato
from ..modelo_interno import DocumentoInterno, LineaInterna
from ..registry import registrar

COLUMNAS_ESPERADAS = [
    "NÚMERO DE DOCUMENTO",
    "TIPO DE COMPROBANTE",
    "CÓDIGO COMPROBANTE",
    "CUENTA CONTABLE",
    "DÉBITO O CRÉDITO",
    "LÍNEA PRODUCTO",
    "GRUPO PRODUCTO",
    "CÓDIGO PRODUCTO",
    "CANTIDAD",
    "LOTE",
    "NIT",
    "CÓDIGO DE LA CIUDAD",
    "DESCRIPCIÓN DE LA SECUENCIA",
]

CODIGOS_PERMITIDOS = {"F": 1, "H": 5, "S": 1, "T": 10}


def _normalizar(s: str) -> str:
    return " ".join(s.upper().split())


def _buscar_columna(encabezados: list[str], esperada: str) -> int:
    """Busca una columna cuyo encabezado CONTENGA el texto esperado. Devuelve índice 1-based o 0."""
    normalizada_esp = _normalizar(esperada)
    for idx, enc in enumerate(encabezados, start=1):
        if normalizada_esp in _normalizar(enc):
            return idx
    return 0


def _a_str(valor) -> str:
    """Convierte un valor de celda a string de forma segura, preservando puntos."""
    if valor is None:
        return ""
    if isinstance(valor, float):
        return f"{valor:.6f}".rstrip("0").rstrip(".")
    return str(valor)


def _a_int_o_str(valor) -> str:
    """Convierte un valor de celda a entero string. Maneja floats (150.0 → 150)."""
    if valor is None:
        return ""
    if isinstance(valor, float):
        return str(int(valor))
    return str(valor).strip()


def _armar_codigo_producto(linea_raw, grupo_raw, codigo_raw) -> str:
    """
    Concatena LÍNEA(3) + GRUPO(4) + CÓDIGO(6) → string de 13 caracteres.
    Ejemplo: 150, 5, 5 → 1500005000005
    """
    linea = _a_int_o_str(linea_raw).strip()
    grupo = _a_int_o_str(grupo_raw).strip()
    codigo = _a_int_o_str(codigo_raw).strip()

    if not linea or not grupo or not codigo:
        raise ValueError("Código de producto incompleto: una o más partes vacías")

    try:
        int(linea)
        int(grupo)
        int(codigo)
    except (ValueError, TypeError):
        raise ValueError(
            f"Código de producto no numérico: línea={linea}, grupo={grupo}, código={codigo}"
        )

    return linea.zfill(3) + grupo.zfill(4) + codigo.zfill(6)


@registrar
class AdaptadorPlantilla(AdaptadorFormato):
    nombre = "PLANTILLA"

    def validar(self, ruta_archivo: Path) -> None:
        if not ruta_archivo.exists():
            raise ValueError(f"Archivo no encontrado: {ruta_archivo}")

        wb = load_workbook(ruta_archivo, read_only=True, data_only=True)

        if "Hoja1" not in wb.sheetnames:
            wb.close()
            raise ValueError(
                "El archivo no contiene la hoja 'Hoja1'. "
                "Estructura esperada del formato: datos en Hoja1."
            )

        ws = wb["Hoja1"]

        # iter_rows funciona aunque max_column/max_row sean None (archivos sin <dimension>)
        encabezados_leidos = []
        for row in ws.iter_rows(min_row=5, max_row=5, values_only=True):
            encabezados_leidos = [str(v).strip() if v is not None else "" for v in row]
            break

        wb.close()

        if not any(encabezados_leidos):
            raise ValueError(
                "El archivo no contiene datos en la fila 5 de 'Hoja1'. "
                "Verifique que el archivo no esté vacío."
            )

        faltantes = []
        for col_esperada in COLUMNAS_ESPERADAS:
            if _buscar_columna(encabezados_leidos, col_esperada) == 0:
                faltantes.append(col_esperada)

        if faltantes:
            raise ValueError(
                f"El archivo no tiene todas las columnas esperadas en la fila 5. "
                f"Faltan: {', '.join(faltantes)}"
            )

    def parse(self, ruta_archivo: Path) -> list[DocumentoInterno]:
        self.validar(ruta_archivo)

        wb = load_workbook(ruta_archivo, read_only=True, data_only=True)
        ws = wb["Hoja1"]

        # Leer fila 5 como encabezados usando iter_rows (funciona sin <dimension>)
        header_row = 5
        encabezados_leidos = []
        for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
            encabezados_leidos = [str(v).strip() if v is not None else "" for v in row]
            break

        if not encabezados_leidos:
            wb.close()
            return []

        col_idx: dict[str, int] = {}
        for esperada in COLUMNAS_ESPERADAS:
            idx = _buscar_columna(encabezados_leidos, esperada)
            if idx:
                col_idx[esperada] = idx

        # Agrupar líneas por documento
        documentos: dict[str, DocumentoInterno] = {}

        def _celda(fila, nombre):
            idx = col_idx.get(nombre, 0)
            if not idx or idx > len(fila):
                return None
            return fila[idx - 1]

        for fila in ws.iter_rows(min_row=header_row + 1, values_only=True):
            num_doc = _a_str(_celda(fila, "NÚMERO DE DOCUMENTO"))
            if not num_doc or not num_doc.strip():
                continue

            tipo_comprobante = _a_str(_celda(fila, "TIPO DE COMPROBANTE")).strip().upper()
            codigo_comprobante = _a_str(_celda(fila, "CÓDIGO COMPROBANTE")).strip()
            cuenta_contable = _a_str(_celda(fila, "CUENTA CONTABLE")).strip()
            debito_credito = _a_str(_celda(fila, "DÉBITO O CRÉDITO")).strip().upper()
            linea_producto = _a_str(_celda(fila, "LÍNEA PRODUCTO")).strip()
            grupo_producto = _a_str(_celda(fila, "GRUPO PRODUCTO")).strip()
            codigo_producto = _a_str(_celda(fila, "CÓDIGO PRODUCTO")).strip()
            cantidad = _celda(fila, "CANTIDAD")
            lote = _a_str(_celda(fila, "LOTE"))
            nit = _a_str(_celda(fila, "NIT")).strip()
            codigo_ciudad = _a_str(_celda(fila, "CÓDIGO DE LA CIUDAD")).strip()
            descripcion = _a_str(_celda(fila, "DESCRIPCIÓN DE LA SECUENCIA")).strip()

            # Criterios de filtro
            if "TRASPORTE" in descripcion.upper():
                continue

            if tipo_comprobante not in CODIGOS_PERMITIDOS:
                continue

            codigo_esperado = CODIGOS_PERMITIDOS[tipo_comprobante]
            try:
                codigo_int = int(float(codigo_comprobante))
            except (ValueError, TypeError):
                continue
            if codigo_int != codigo_esperado:
                continue

            if not cuenta_contable.startswith("14"):
                continue

            # T+10 (traslados) acepta tanto D como C; los demás solo C
            es_traslado = tipo_comprobante == "T" and codigo_int == 10
            if not es_traslado and debito_credito != "C":
                continue

            cantidad_dec = Decimal("0")
            if cantidad is not None:
                try:
                    cantidad_dec = Decimal(str(cantidad))
                except Exception:
                    pass

            # producto_codigo — armar concatenando las 3 partes
            try:
                codigo_producto_final = _armar_codigo_producto(
                    linea_producto, grupo_producto, codigo_producto
                )
            except ValueError:
                codigo_producto_final = ""

            linea = LineaInterna(
                producto_codigo=codigo_producto_final,
                lote_raw=lote,
                cantidad_origen=cantidad_dec,
                descripcion_origen=descripcion,
            )

            if num_doc not in documentos:
                documentos[num_doc] = DocumentoInterno(
                    factura=num_doc,
                    nit=nit,
                    codigo_ciudad=codigo_ciudad,
                    tipo_comprobante=tipo_comprobante,
                )
            documentos[num_doc].lineas.append(linea)

        wb.close()
        return list(documentos.values())
