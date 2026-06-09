from pathlib import Path
from datetime import date

from django.test import TestCase, override_settings
from django.urls import reverse
from django.contrib.auth.models import User, Group
from openpyxl import Workbook

from cortes.models import Corte
from productos.models import Producto


ENCABEZADOS = [
    "NÚMERO DE DOCUMENTO",
    "TIPO DE COMPROBANTE",
    "CÓDIGO COMPROBANTE",
    "CUENTA CONTABLE",
    "DÉBITO O CRÉDITO",
    "LÍNEA PRODUCTO",
    "GRUPO PRODUCTO",
    "CÓDIGO PRODUCTO",
    "CANTIDAD",
    "LOTE",
    "NIT",
    "CÓDIGO DE LA CIUDAD",
    "DESCRIPCIÓN DE LA SECUENCIA",
]


def _crear_excel_valido(ruta: Path, factura="DOC001") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    for col, encabezado in enumerate(ENCABEZADOS, start=1):
        ws.cell(row=5, column=col, value=encabezado)
    fila = [factura, "F", "1", "143505", "C", "150", "0005", "000005",
            "10", "15F22", "800000", "11001", "Desc"]
    for j, valor in enumerate(fila, start=1):
        ws.cell(row=6, column=j, value=valor)
    wb.save(ruta)


def _crear_excel_malformado(ruta: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "OtraHoja"
    ws.cell(row=1, column=1, value="No soy Hoja1")
    wb.save(ruta)


@override_settings(MEDIA_ROOT="/tmp/test_media")
class VistaCargarTest(TestCase):
    def setUp(self):
        Path("/tmp/test_media/uploads").mkdir(parents=True, exist_ok=True)

        self.facturacion_user = User.objects.create_user(username="facturacion", password="test")
        self.consulta_user = User.objects.create_user(username="consulta", password="test")
        self.admin_user = User.objects.create_user(username="admin_test", password="test")

        grupo_facturacion, _ = Group.objects.get_or_create(name="facturacion")
        grupo_consulta, _ = Group.objects.get_or_create(name="consulta")
        grupo_admin, _ = Group.objects.get_or_create(name="admin")

        self.facturacion_user.groups.add(grupo_facturacion)
        self.consulta_user.groups.add(grupo_consulta)
        self.admin_user.groups.add(grupo_admin)

        Producto.objects.create(
            producto="1500005000005",
            referencia="REF1",
            descripcion="DESC1",
            unidad_empaque=4,
            activo=True,
        )

    def test_login_obligatorio(self):
        response = self.client.get(reverse("cargar_corte"))
        self.assertRedirects(response, "/admin/login/?next=/cortes/cargar/")

    def test_consulta_no_puede_cargar(self):
        self.client.login(username="consulta", password="test")
        response = self.client.get(reverse("cargar_corte"))
        self.assertEqual(response.status_code, 403)

    def test_facturacion_puede_ver_formulario(self):
        self.client.login(username="facturacion", password="test")
        response = self.client.get(reverse("cargar_corte"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cargar nuevo corte")

    def test_carga_exitosa(self):
        self.client.login(username="facturacion", password="test")

        ruta = Path("/tmp/test_carga_exitosa.xlsx")
        _crear_excel_valido(ruta)

        with open(ruta, "rb") as f:
            response = self.client.post(
                reverse("cargar_corte"),
                {
                    "archivo": f,
                    "formato_origen": "PLANTILLA",
                    "numero_corte": "2",
                },
            )

        ruta.unlink(missing_ok=True)

        self.assertEqual(response.status_code, 302)
        corte = Corte.objects.first()
        self.assertIsNotNone(corte)
        self.assertEqual(corte.estado, "en_revision")
        self.assertGreater(corte.documentos.count(), 0)

    def test_archivo_duplicado(self):
        self.client.login(username="facturacion", password="test")

        ruta = Path("/tmp/test_dup.xlsx")
        _crear_excel_valido(ruta)

        with open(ruta, "rb") as f:
            self.client.post(
                reverse("cargar_corte"),
                {"archivo": f, "formato_origen": "PLANTILLA", "numero_corte": "1"},
            )

        with open(ruta, "rb") as f:
            response = self.client.post(
                reverse("cargar_corte"),
                {"archivo": f, "formato_origen": "PLANTILLA", "numero_corte": "2"},
            )

        ruta.unlink(missing_ok=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ya existe un corte con este archivo")

    def test_archivo_malformado(self):
        self.client.login(username="facturacion", password="test")

        ruta = Path("/tmp/test_mal.xlsx")
        _crear_excel_malformado(ruta)

        with open(ruta, "rb") as f:
            response = self.client.post(
                reverse("cargar_corte"),
                {"archivo": f, "formato_origen": "PLANTILLA", "numero_corte": "1"},
            )

        ruta.unlink(missing_ok=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Hoja1")
